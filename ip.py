import urllib.request
import IPy
import re
from openpyxl import load_workbook

#通过tb查询IP归属地
def ipsearch(ip):
    ipsurl = 'http://ip.taobao.com/service/getIpInfo.php?ip='
    url = ipsurl+ip
    res = eval(str(urllib.request.urlopen(url).read(),encoding='utf-8'))['data']
    country = res['country']
    region = res['region']
    city = res['city']
    return country,region,city
#判断IP类型
def iptype(ip):
    try:
        ip_type = IPy.IP(ip).iptype()
    except:
        ip_type = 'UNKNOWN'
    return ip_type

wb = load_workbook("D:/wl.xlsx")
sheet1 = wb["wl"]
nrows1 = sheet1.max_row

col = int(input("外部IP所在列数："))
for i in range(nrows1):
    try:
        ip = re.split(':',sheet1.cell(i+2,col).value)[0]
        ip_type = iptype(ip)
        if ip_type == 'PUBLIC':
            sheet1.cell(row=i + 2, column=8, value=str(ipsearch(ip)))
        else:
            i+=1

    except:
        i+=1
wb.save('D:/wlres.xlsx')











